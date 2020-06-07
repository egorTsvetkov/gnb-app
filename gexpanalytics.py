import math
from decimal import *
import gexpmath as gm
import numpy as np



def integral_function_whole(k, x, params):
	u, p, theta = params
	numerator = abs(u) * Decimal(math.exp(-((x / theta) ** u))) * (x ** (u * p - 1)) 
	denumenator = theta ** (u * p) * Decimal(math.gamma(p))
	poiss_part = 0
	if lamb == 0:
		poiss_part = Decimal(1 / math.factorial(k))
	else:
		poiss_part = Decimal(math.exp(-x)) * (x ** k) / math.factorial(k)

	return poisson
def integral_function(k, x, params):
	return gm.poisson(k, x) * gm.generalized_gamma(x, params)

def gauss5(a, b, k, params):
	
	ksi = [Decimal(-math.sqrt((35 + 2 * math.sqrt(70)) / 63)), 
		   Decimal(-math.sqrt((35 - 2 * math.sqrt(70)) / 63)), 
		   Decimal(0), 
		   Decimal(math.sqrt((35 - 2 * math.sqrt(70)) / 63)), 
		   Decimal(math.sqrt((35 + 2 * math.sqrt(70)) / 63))]
	
	gamma = [Decimal((322 + 13 * math.sqrt(70)) / 900), 
			 Decimal((322 - 13 * math.sqrt(70)) / 900), 
			 Decimal(128 / 225), 
			 Decimal((322 - 13 * math.sqrt(70)) / 900), 
			 Decimal((322 + 13 * math.sqrt(70)) / 900)]

	x = []
	c = []

	for i in range(5):
		
		x.append((a + b) / 2 + (b - a) * ksi[i] / 2)
		c.append(gamma[i] * (b - a) / 2)

	integral = 0

	for i in range(5):
		integral += integral_function(k, Decimal(x[i]), params) * Decimal(c[i])

	return integral

def simpson(a, b, k, params):
	coef = Decimal((b - a) / 6)

	return coef * (integral_function(k, a, params) + 4 * integral_function(k, (a + b) / 2, params) + integral_function(k, b, params))


def trapezoidal(a, b, k, params):
	return (b - a) / 2 * (integral_function(k, a, params) + integral_function(k, b, params))


def numeric_integration(k, B, n, u, p, theta):

	
	h = B / n
	params = (u, p, theta)

	print('params =', params)
	print("Probability that N =", k)
	print("-" * 30)
	gnb_value = gm.GNB_Probability(k, u, p, theta)
	print("Value by Gexp function =", gnb_value)
	print("-" * 30)


	# simpson_integral_value = 0
	# for i in range(n):
	# 	simpson_integral_value += simpson(Decimal(i * h), Decimal((i + 1) * h), k, params)

	# print("Simpson method integral =", simpson_integral_value)
	# print("-" * 30)

	# trapezoidal_integral_value = 0
	# for i in range(n):
	# 	trapezoidal_integral_value += trapezoidal(Decimal(i * h), Decimal((i + 1) * h), k, params)

	# print("Trapezoidal method integral =", trapezoidal_integral_value)
	# print("-" * 30)

	gauss_integral_value = 0
	for i in range(n):
		curvalue = gauss5(Decimal(i * h), Decimal((i + 1) * h), k, params)
		if (i % 1000 == 0):
			print(i)
			print(curvalue)
			print(gauss_integral_value)

		gauss_integral_value += curvalue

	print("Gauss 5 points method integral =", gauss_integral_value)
	print("-" * 30)

	trapezoidal_integral_value = 0
	for i in range(n):
		eps = 0
		if i == 0:
			eps = 0.0000001
		curvalue = trapezoidal(Decimal(i * h + Decimal(eps)), Decimal((i + 1) * h), k, params)
		if (i % 1000 == 0):
			print(i)
			print(curvalue)
			print(trapezoidal_integral_value)

		trapezoidal_integral_value += curvalue

	print("Trapezoidal method integral =", gauss_integral_value)
	print("-" * 30)
	print()
	return gnb_value, gauss_integral_value, trapezoidal_integral_value


B = Decimal('100')
n = 1000000
u = [Decimal('0.4'), Decimal('0.7'), Decimal('0.9'), Decimal('1.1'), Decimal('1.3'), Decimal('1.6')]
p = [Decimal('0.9'), Decimal('1.5'), Decimal('1.5'), Decimal('1.5'), Decimal('2.0'), Decimal('3.0')]
theta = [Decimal('0.5'), Decimal('2.0'), Decimal('2.0'), Decimal('2'), Decimal('3.0'), Decimal('1.5')]

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

for i in range(len(u)):
	ws = wb.create_sheet("Sheet" + str(i))
	ws['A1'] = 'u'
	ws['B1'] = 'p'
	ws['C1'] = 'theta'
	ws['A2'] = str(u[i])
	ws['B2'] = str(p[i])
	ws['C2'] = str(theta[i])
	ws['A3'] = "Гамма-экспоненциальная функция"
	ws['B3'] = "Пятиточечный метод Гаусса"
	ws['C3'] = "Метод трапеций"
	for j in range (10):
		gnb_value, gauss_integral_value, trapezoidal_integral_value = numeric_integration(j, B, n, u[i], p[i], theta[i])
		ws['A' + str(j + 4)] = gnb_value
		ws['B' + str(j + 4)] = gauss_integral_value
		ws['C' + str(j + 4)] = trapezoidal_integral_value

wb.save('Сравнение_с_численными_методами.xls')
# gm.generalized_gamma(Decimal(0), (Decimal('0.1'), Decimal('0.2'), Decimal('0.2')))
# from openpyxl import Workbook

# probrange = 10
# u_idx = 0
# import itertools
# for u in itertools.chain(np.arange(0.1, 1, 0.1), np.arange(1.1, 3, 0.1)):
# 	wb = Workbook()
# 	print('u =', u)
	
# 	for p in np.arange(0.2, 4, 0.2):
# 		ws = wb.create_sheet("Sheet p = " + str(p))
# 		print('p =', p)
# 		currentrow = 1
# 		for theta in np.arange(0.2, 4, 0.2):
# 			ws['A' + str(currentrow)] = 'u'
# 			ws['B' + str(currentrow)] = u
# 			ws['A' + str(currentrow + 1)] = 'p'
# 			ws['B' + str(currentrow + 1)] = p
# 			ws['A' + str(currentrow + 2)] = 'alpha'
# 			ws['B' + str(currentrow + 2)] = theta
# 			ws['A' + str(currentrow + 3)] = 'k'
# 			ws['B' + str(currentrow + 3)] = 'P(N = k), гамма-экспоненциальная функция'
# 			# ws['C' + str(currentrow + 3)] = 'P(N = k), пятиточечный метод гаусса'

# 			currentrow += 4
# 			tmp = []
# 			for j in range(probrange):
# 				ws['A' + str(currentrow)] = j
# 				# ws['B' + str(currentrow)] = numeric_integration(j, B, n, u[i], p[i], theta[i])
# 				ws['B' + str(currentrow)] = gm.GNB_Probability(j, u, p, theta)
# 				currentrow += 1
# 			currentrow += 1

# 			wb.save('GNB_u=' + str(u) + '.xls')
			

# 		print('New sheet')
		
		

	




# wb.save('GNB.xlsx')

# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active

# umin = Decimal(0.1)
# umax = Decimal(1)
# ustep = Decimal(0.1)

# pmin = Decimal(0.2)
# pmax = Decimal(3.2)
# pstep = Decimal(0.2)

# thetamin = Decimal(0.2)
# thetamax = Decimal(3.2)
# thetastep = Decimal(0.2)

# probrange = 10

# currentrow = 1
# ws['C' + str(currentrow)] = ''
# for u in np.arange(umin, umax, ustep):
# 	for p in np.arange(pmin, pmax, pstep):
# 		for theta in np.arange(thetamin, thetamax, thetastep):

# 			ws['A' + str(currentrow)] = 'u'
# 			ws['B' + str(currentrow)] = u
# 			ws['A' + str(currentrow + 1)] = 'p'
# 			ws['B' + str(currentrow + 1)] = p
# 			ws['A' + str(currentrow + 2)] = 'alpha'
# 			ws['B' + str(currentrow + 2)] = theta
# 			ws['A' + str(currentrow + 3)] = 'k'
# 			ws['B' + str(currentrow + 3)] = 'P(N = k), гамма-экспоненциальная функция'
# 			ws['C' + str(currentrow + 3)] = 'P(N = k), пятиточечный метод гаусса'

# 			currentrow += 4

# 			for i in range(probrange):

# 				ws['A' + str(currentrow)] = i
# 				ws['B' + str(currentrow)] = numeric_integration(i, B, n, u, p, theta)
# 				ws['C' + str(currentrow)] = gm.GNB_Probability(i, u, p, theta)
# 				currentrow += 1

# 			currentrow += 2

# umin = Decimal(1.1)
# umax = Decimal(2)
# ustep = Decimal(0.1)

# for u in np.arange(umin, umax, ustep):
# 	for p in np.arange(pmin, pmax, pstep):
# 		for theta in np.arange(thetamin, thetamax, thetastep):

# 			ws['A' + str(currentrow)] = 'u'
# 			ws['B' + str(currentrow)] = u

# 			ws['A' + str(currentrow + 1)] = 'p'
# 			ws['B' + str(currentrow + 1)] = p
# 			ws['A' + str(currentrow + 2)] = 'alpha'
# 			ws['B' + str(currentrow + 2)] = theta
# 			ws['A' + str(currentrow + 3)] = 'k'
# 			ws['B' + str(currentrow + 3)] = 'P(N = k), гамма-экспоненциальная функция'
# 			ws['C' + str(currentrow + 3)] = 'P(N = k), пятиточечный метод гаусса'

# 			currentrow += 4

# 			for i in range(probrange):

# 				ws['A' + str(currentrow)] = i
# 				ws['B' + str(currentrow)] = numeric_integration(i, B, n, u, p, theta)
# 				ws['C' + str(currentrow)] = gm.GNB_Probability(i, u, p, theta)
# 				currentrow += 1

# 			currentrow += 2

# wb.save('GNB.xlsx')


