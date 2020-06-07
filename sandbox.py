from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws['A1'] = 'x'

ws = wb.create_sheet("Mysheet")
ws['A1'] = 'x'
wb.save('test.xls')